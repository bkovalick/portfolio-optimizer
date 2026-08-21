import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from io import BytesIO
from models.experiment import Experiment


logger = logging.getLogger(__name__)

def deserialize_series(data) -> pd.Series:
    """Deserialize a series from JSON round-trip (handles {index, values} format and plain dicts)."""
    if isinstance(data, pd.Series):
        return data
    if isinstance(data, dict) and "index" in data and "values" in data:
        return pd.Series(data["values"], index=data["index"])
    if isinstance(data, dict):
        return pd.Series(data)
    return pd.Series(data)

def deserialize_dataframe(data) -> pd.DataFrame:
    """Deserialize a DataFrame from JSON round-trip."""
    if isinstance(data, pd.DataFrame):
        return data.copy()
    if isinstance(data, dict) and "index" in data and "columns" in data and "values" in data:
        return pd.DataFrame(data["values"], index=data["index"], columns=data["columns"])
    if isinstance(data, dict):
        try:
            return pd.DataFrame.from_dict(data, orient="index")
        except Exception:
            return pd.DataFrame(data)
    return pd.DataFrame(data)

class ExcelGenerator:
    def __init__(self, experiment: Experiment, buffer: BytesIO):
        self.experiment = experiment
        cfg = experiment.market_config
        if not isinstance(cfg, dict):
            cfg = cfg.model_dump() if hasattr(cfg, "model_dump") else (cfg.dict() if hasattr(cfg, "dict") else vars(cfg))
        self.config = cfg
        self.buffer = buffer

    def generate_report(self):
        ic_results = self.aggregate_ic_series()
        results = self.aggregate_performance_metrics()
        results.update(ic_results)

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        if "summary" in results and results["summary"] is not None:
            summary_ws = wb.create_sheet(title="Summary")
            for r_idx, row in enumerate(dataframe_to_rows(results["summary"], header=True, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    summary_ws.cell(row=r_idx, column=c_idx, value=value)

        if "time_series" in results and results["time_series"] is not None:
            ts_ws = wb.create_sheet(title="Time Series")
            for r_idx, row in enumerate(dataframe_to_rows(results["time_series"], header=True, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)

        if "rolling_time_series" in results and results["rolling_time_series"] is not None:
            ts_ws = wb.create_sheet(title="Rolling Time Series")
            for r_idx, row in enumerate(dataframe_to_rows(results["rolling_time_series"], header=True, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)

        if "ic_summary" in results and results["ic_summary"] is not None:
            ts_ws = wb.create_sheet(title="IC Summary")
            for r_idx, row in enumerate(dataframe_to_rows(results["ic_summary"], header=True, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)               

        if "ic_series" in results and results["ic_series"] is not None:
            ts_ws = wb.create_sheet(title="IC Series")
            for r_idx, row in enumerate(dataframe_to_rows(results["ic_series"], header=True, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)

        if "factor_regression" in results and results["factor_regression"] is not None:
            reg_ws = wb.create_sheet(title="Factor Regression")
            reg_df = results["factor_regression"]
            for _, rec in reg_df.iterrows():
                reg_ws.append([rec["strategy"]])
                for k, v in rec["regression_summary"].items():
                    line = f"{k}: {v}"
                    reg_ws.append([line])
                reg_ws.append([]) 

        wb.save(self.buffer)
        self.buffer.seek(0)

    def aggregate_ic_series(self):
        ic_summary_rows = []
        ic_statistics_agg_df = []
        regression_rows = []

        for strategy_run in self.experiment.strategy_runs:
            if strategy_run.monitoring_stats is None:
                continue

            stats = strategy_run.monitoring_stats
            strategy_name = strategy_run.strategy_name

            if stats.ic_summary is not None:
                row = {"strategy": strategy_name}
                for k, v in stats.ic_summary.items():
                    if isinstance(v, (pd.Series, pd.DataFrame)):
                        continue
                    row[k] = v
                ic_summary_rows.append(row)

            if stats.ic_statistics is not None:
                ic_statistics_df = deserialize_dataframe(stats.ic_statistics).T
                if len(ic_statistics_df) > 1:
                    ic_statistics_df.insert(0, "Date", pd.to_datetime(ic_statistics_df.index))
                    ic_statistics_df = ic_statistics_df.reset_index(drop=True)
                    ic_statistics_df.insert(1, "Strategy", strategy_name)
                    ic_statistics_df = ic_statistics_df.rename(columns={0: "IC_Series"})
                ic_statistics_agg_df.append(ic_statistics_df)

            if stats.regression_summary is not None:
                regression_rows.append({"strategy": strategy_name, "regression_summary": stats.regression_summary})

        ic_summary_df = pd.DataFrame(ic_summary_rows) if ic_summary_rows else None
        ic_statistics_agg_df = pd.concat(ic_statistics_agg_df, axis=0, ignore_index=True) if ic_statistics_agg_df else None
        regression_df = pd.DataFrame(regression_rows) if regression_rows else None
        return {
            "ic_summary": ic_summary_df,
            "ic_series": ic_statistics_agg_df,
            "factor_regression": regression_df,
        }

    def aggregate_performance_metrics(self):
        """Aggregate performance metrics from multiple strategies into summary and time series DataFrames."""
        summary_rows = []
        portfolio_dfs = []
        rolling_dfs = []

        for strategy_run in self.experiment.strategy_runs:
            strategy_name = strategy_run.strategy_name

            # Summary
            row = {"strategy": strategy_name}
            for k, v in strategy_run.result.summary.items():
                if isinstance(v, (pd.Series, pd.DataFrame)):
                    continue
                row[k] = v
            summary_rows.append(row)

            # Time series
            series = strategy_run.result.series
            if "portfolio_weights" in series:
                try:
                    weights_df = deserialize_dataframe(series["portfolio_weights"])
                    wealth_series = deserialize_series(series["portfolio_wealth_factors"])
                    returns_series = deserialize_series(series["portfolio_returns"])
                    turnover_series = deserialize_series(series["portfolio_turnover"])
                    trades_series = deserialize_series(series["portfolio_trades"])
                    benchmark_weights_series = deserialize_series(series["benchmark_wealth_factors"])
                    benchmark_returns_series = deserialize_series(series["benchmark_returns"])
                    weights_df = weights_df.reset_index(drop=True)
                    min_len = min(
                        len(weights_df), len(wealth_series),
                        len(returns_series), len(turnover_series), len(trades_series)
                    )
                    weights_df = weights_df.iloc[:min_len].copy()
                    date_idx = pd.to_datetime(wealth_series.index[:min_len])
                    # Align benchmark to the strategy's exact date index so every
                    # strategy row carries the same benchmark value for the same date.
                    bm_wealth_aligned = benchmark_weights_series.reindex(wealth_series.index[:min_len])
                    bm_returns_aligned = benchmark_returns_series.reindex(wealth_series.index[:min_len])
                    weights_df.insert(0, "Date", date_idx)
                    weights_df.insert(1, "Strategy", strategy_name)
                    weights_df.insert(2, "StrategyWealthFactors", wealth_series.values[:min_len])
                    weights_df.insert(3, "BenchmarkWealthFactors", bm_wealth_aligned.values)
                    weights_df.insert(4, "PortfolioReturns", returns_series.values[:min_len])
                    weights_df.insert(5, "BenchmarkReturns", bm_returns_aligned.values)
                    weights_df.insert(6, "PortfolioTurnover", turnover_series.values[:min_len])
                    weights_df.insert(7, "PortfolioTrades", trades_series.values[:min_len])
                    portfolio_dfs.append(weights_df)
                except Exception as e:
                    logger.warning("Could not build time series for %s: %s", strategy_name, e)

            # Rolling time series
            if "rolling_returns" in series:
                try:
                    rr = deserialize_series(series["rolling_returns"])
                    rolling_df = pd.DataFrame({
                        "Date": pd.to_datetime(rr.index),
                        "Strategy": strategy_name,
                        "RollingReturns": rr.values,
                        "RollingVolatility": deserialize_series(series["rolling_volatility"]).values,
                        "RollingSharpe": deserialize_series(series["rolling_sharpe_ratio"]).values,
                        "RollingDrawdown": deserialize_series(series["rolling_drawdown"]).values,
                        "RollingTurnover": deserialize_series(series["rolling_turnover"]).values,
                    })
                    rolling_dfs.append(rolling_df)
                except Exception as e:
                    logger.warning("Could not build rolling series for %s: %s", strategy_name, e)

        summary_df = pd.DataFrame(summary_rows)
        portfolio_metrics_df = pd.concat(portfolio_dfs, axis=0, ignore_index=True) if portfolio_dfs else None
        rolling_metrics_df = pd.concat(rolling_dfs, axis=0, ignore_index=True) if rolling_dfs else None

        return {
            "summary": summary_df,
            "time_series": portfolio_metrics_df,
            "rolling_time_series": rolling_metrics_df
        }